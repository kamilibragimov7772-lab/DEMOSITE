"""Generate real seed JSON for TERMY demo from Master Knowledge Base xlsx and vault notes."""
import sys, os, json, re, unicodedata
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from datetime import datetime

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
KB_ROOT = os.path.join(ROOT, 'TERMY_Knowledge_Project')
MKB = os.path.join(KB_ROOT, '05_Product_Docs', 'TERMY_Master_Knowledge_Base_v1_1.xlsx')
OUT = os.path.join(ROOT, 'demo_build', 'src', 'data', 'real')
os.makedirs(OUT, exist_ok=True)


def iso(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    return str(v)


def write_json(name, data):
    p = os.path.join(OUT, name)
    with open(p, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f'  wrote {name}: {len(data)} records')


wb = load_workbook(MKB, read_only=True, data_only=True)


def sheet_rows(sheet_name):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [h for h in rows[0]]
    return headers, rows[1:]


def parse_list(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    return [x.strip() for x in re.split(r'[;,]', s) if x.strip()]


# ===== PL01 Product Lines =====
headers, rows = sheet_rows('PL01 Product Lines')
product_lines = []
for row in rows:
    if row[0] is None:
        continue
    d = dict(zip(headers, row))
    product_lines.append({
        'id': d.get('record_id'),
        'name_canonical': d.get('line_name_canonical'),
        'name_variants': parse_list(d.get('line_name_variants')),
        'category': d.get('category'),
        'status': d.get('status'),
        'description': d.get('description'),
        'standard_reference': d.get('standard_reference'),
        'construction_summary': d.get('construction_summary'),
        'source_reference': d.get('source_reference'),
        'confidence_status': d.get('confidence_status'),
        'evidence_type': parse_list(d.get('evidence_type')),
        'conflict_flag': bool(d.get('conflict_flag')) if d.get('conflict_flag') is not None else False,
        'notes': d.get('notes'),
        'created_date': iso(d.get('created_date')),
        'last_updated': iso(d.get('last_updated')),
    })
write_json('product_lines.json', product_lines)

# ===== PV01 SKUs =====
headers, rows = sheet_rows('PV01 Product Variants')
skus = []

def detect_sku_source(product_line_id, variant_name):
    # PL-006 Термошопперы = purchased (Китай), PL-010 Термокружки = purchased (WUYI TOCHISO)
    if product_line_id in ('PL-006', 'PL-010'):
        return 'purchased'
    return 'production'

for row in rows:
    if row[0] is None:
        continue
    d = dict(zip(headers, row))
    pl_id = d.get('product_line_id')
    vname = d.get('variant_name') or ''
    skus.append({
        'id': d.get('record_id'),
        'product_line_id': pl_id,
        'variant_name': vname,
        'sub_series': d.get('sub_series'),
        'size_descriptor': d.get('size_descriptor'),
        'dimensions_mm': d.get('dimensions_mm'),
        'volume_liters': d.get('volume_liters'),
        'load_capacity_kg': d.get('load_capacity_kg'),
        'article_number': d.get('article_number'),
        'ean_13': str(d.get('ean_13')) if d.get('ean_13') else None,
        'pack_quantity': d.get('pack_quantity'),
        'handle_type': d.get('handle_type'),
        'closure_type': d.get('closure_type'),
        'color_options': d.get('color_options'),
        'sku_status': 'active',
        'sku_source': detect_sku_source(pl_id, vname),
        'confidence_status': d.get('confidence_status') or 'confirmed',
        'conflict_flag': bool(d.get('conflict_flag')) if d.get('conflict_flag') is not None else False,
        'source_reference': d.get('source_reference'),
        'notes': d.get('notes'),
    })
write_json('skus.json', skus)

# ===== TS01 Tech Specs =====
headers, rows = sheet_rows('TS01 Technical Specs')
tech_specs = []
for idx, row in enumerate(rows, 1):
    if row[0] is None:
        continue
    d = dict(zip(headers, row))
    tech_specs.append({
        'id': d.get('record_id') or f'TS-{idx:03}',
        'sku_id': d.get('sku_id') or d.get('variant_id'),
        'layer_count': d.get('layer_count'),
        'layer_composition_text': d.get('layer_composition_text') or d.get('composition'),
        'mat_meta_thickness_um': d.get('outer_layer_met_thickness') or d.get('mat_meta_thickness_um'),
        'vpe_thickness_um': d.get('vpe_thickness') or d.get('vpe_thickness_um'),
        'vpp_thickness_cm': d.get('vpp_thickness') or d.get('vpp_thickness_cm'),
        'pvd_thickness_um': d.get('inner_layer_pvd_thickness') or d.get('pvd_thickness_um'),
        'bopp_thickness_um': d.get('bopp_thickness_um'),
        'spanbond_thickness_um': d.get('spanbond_thickness_um'),
        'holding_time_hours': d.get('temperature_hold_hours') or d.get('holding_time_hours'),
        'holding_time_description': d.get('holding_time_description') or d.get('temperature_hold_hours_pvd') or d.get('temperature_hold_hours_met'),
        'temperature_range': d.get('temperature_range'),
        'bubble_diameter': d.get('bubble_diameter'),
        'notes': d.get('notes'),
        'confidence_status': d.get('confidence_status') or 'confirmed',
        'conflict_flag': bool(d.get('conflict_flag')) if d.get('conflict_flag') is not None else False,
    })
write_json('tech_specs.json', tech_specs)

# ===== PM01 Printing Methods =====
headers, rows = sheet_rows('PM01 Printing Methods')
printing_methods = []
for idx, row in enumerate(rows, 1):
    if row[0] is None:
        continue
    d = dict(zip(headers, row))
    printing_methods.append({
        'id': d.get('record_id') or f'PM-{idx:02}',
        'name': d.get('method_name') or d.get('name'),
        'surface_types': parse_list(d.get('surface_types')),
        'max_colors': d.get('max_colors'),
        'min_order': d.get('min_order') or d.get('moq'),
        'optimal_order': d.get('optimal_order'),
        'lead_time_weeks_min': d.get('lead_time_weeks_min'),
        'lead_time_weeks_max': d.get('lead_time_weeks_max'),
        'gradient_allowed': bool(d.get('gradient_allowed')) if d.get('gradient_allowed') is not None else None,
        'edge_printing_allowed': bool(d.get('edge_printing_allowed')) if d.get('edge_printing_allowed') is not None else None,
        'default_print_area': d.get('default_print_area'),
        'contractor': d.get('contractor'),
        'stability_note': d.get('stability_note') or d.get('notes'),
    })
# fallback, if PM01 empty — hardcode from Виды печати.md
if not printing_methods:
    printing_methods = [
        {'id': 'PM-01', 'name': 'Шелкография', 'surface_types': ['ПВД'], 'max_colors': 7, 'min_order': 100, 'optimal_order': 500,
         'lead_time_weeks_min': 3, 'lead_time_weeks_max': 5, 'gradient_allowed': False, 'edge_printing_allowed': False,
         'default_print_area': '30×30 см', 'contractor': 'Лаир',
         'stability_note': 'Отмечена нестабильность цветов (красный→бордовый, белый полупрозрачный). Идёт поиск альтернативы.'},
        {'id': 'PM-02', 'name': 'Флексография', 'surface_types': ['ПВД', 'металлизированная плёнка'], 'max_colors': 10, 'min_order': 10000, 'optimal_order': None,
         'lead_time_weeks_min': 4, 'lead_time_weeks_max': 5, 'gradient_allowed': True, 'edge_printing_allowed': True,
         'default_print_area': None, 'contractor': 'Партнёр-Медиа',
         'stability_note': 'ФЛФ силиконовая. Ресурс 1 000 000 оттисков / 1,5 года / 6 приладок.'},
    ]
write_json('printing_methods.json', printing_methods)

# ===== PC01 Print Constraints =====
headers, rows = sheet_rows('PC01 Print Constraints')
print_constraints = []
for idx, row in enumerate(rows, 1):
    if row[0] is None:
        continue
    d = dict(zip(headers, row))
    print_constraints.append({
        'id': d.get('record_id') or f'PC-{idx:03}',
        'parameter': d.get('parameter'),
        'value': str(d.get('value')) if d.get('value') is not None else '',
        'applies_to_method_id': d.get('applies_to_method_id'),
        'rule_type': d.get('rule_type') or 'mandatory',
        'description': d.get('description'),
    })
write_json('print_constraints.json', print_constraints)

# ===== CP01 Color Palette =====
headers, rows = sheet_rows('CP01 Color Palette')
palette = []
for idx, row in enumerate(rows, 1):
    if row[0] is None:
        continue
    d = dict(zip(headers, row))
    palette.append({
        'id': d.get('record_id') or f'CP-{idx:03}',
        'pantone_code': d.get('pantone_code') or d.get('code'),
        'cmyk_code': d.get('cmyk_code'),
        'name_ru': d.get('name_ru') or d.get('color_name'),
        'available_for_method_id': d.get('available_for_method_id'),
    })
write_json('color_palette.json', palette)

# ===== CL01 Claims =====
headers, rows = sheet_rows('CL01 Claims')
claims = []
for idx, row in enumerate(rows, 1):
    if row[0] is None:
        continue
    d = dict(zip(headers, row))
    claims.append({
        'id': d.get('record_id') or f'CL-{idx:03}',
        'text_ru': d.get('claim_text') or d.get('text') or d.get('text_ru'),
        'segment': d.get('segment'),
        'evidence_level': d.get('evidence_level') or 'internal',
        'public_allowed': bool(d.get('public_allowed')) if d.get('public_allowed') is not None else False,
        'linked_to': [],
    })
write_json('claims.json', claims)

# ===== Regulatory documents — build from B04 Certificates + B03 Trademark + hard facts =====
try:
    headers, rows = sheet_rows('B04 Certificates')
    docs = []
    for idx, row in enumerate(rows, 1):
        if row[0] is None:
            continue
        d = dict(zip(headers, row))
        docs.append({
            'id': d.get('record_id') or f'DOC-{idx:03}',
            'doc_type': d.get('doc_type') or 'certificate',
            'number': str(d.get('number') or d.get('certificate_number') or '').strip(),
            'title': d.get('title') or d.get('name') or '',
            'issuing_authority': d.get('issuing_authority'),
            'issued_date': iso(d.get('issued_date')),
            'expiry_date': iso(d.get('expiry_date')),
            'status': d.get('status') or 'valid',
            'file_path': d.get('file_path'),
            'public_allowed': bool(d.get('public_allowed')) if d.get('public_allowed') is not None else True,
            'applicable_sku_ids': [],
            'notes': d.get('notes'),
            'source_reference': d.get('source_reference'),
        })
except Exception:
    docs = []

# Enrich / override with canonical docs from Патенты и сертификаты.md
canonical = [
    {'id': 'DOC-DEC-902483', 'doc_type': 'declaration', 'number': '902483',
     'title': 'Декларация ЕАЭС — ТР ТС 005/2011 (упаковка)', 'issuing_authority': 'ЕАЭС',
     'issued_date': '2022-09-03', 'expiry_date': '2032-08-10', 'status': 'valid',
     'file_path': '/documents/902483.eod.pdf', 'public_allowed': True,
     'applicable_sku_ids': [], 'notes': 'Заявка №2022714614, зарегистрирована 10.08.2022', 'source_reference': '01_Obsidian_Vault/06-Документы/Патенты и сертификаты.md'},
    {'id': 'DOC-DEC-1158277', 'doc_type': 'declaration', 'number': '1158277',
     'title': 'Декларация ЕАЭС — ТР ТС 005/2011', 'issuing_authority': 'ЕАЭС',
     'issued_date': '2025-01-17', 'expiry_date': '2034-12-13', 'status': 'valid',
     'file_path': '/documents/1158277.eod.pdf', 'public_allowed': True,
     'applicable_sku_ids': [], 'notes': 'Заявка №2024803234, зарегистрирована 13.12.2024', 'source_reference': 'Патенты и сертификаты.md'},
    {'id': 'DOC-DEC-1158278', 'doc_type': 'declaration', 'number': '1158278',
     'title': 'Декларация ЕАЭС — ТР ТС 005/2011', 'issuing_authority': 'ЕАЭС',
     'issued_date': '2025-01-17', 'expiry_date': '2034-12-13', 'status': 'valid',
     'file_path': '/documents/1158278.eod.pdf', 'public_allowed': True,
     'applicable_sku_ids': [], 'notes': 'Заявка №2024803238', 'source_reference': 'Патенты и сертификаты.md'},
    {'id': 'DOC-ISO-9001', 'doc_type': 'certificate', 'number': 'TBD',
     'title': 'ISO 9001', 'issuing_authority': None,
     'issued_date': None, 'expiry_date': None, 'status': 'valid',
     'file_path': '/documents/sertifikat.pdf', 'public_allowed': True,
     'applicable_sku_ids': [], 'notes': 'R03 Master KB: номер сертификата уточняется.',
     'source_reference': 'Патенты и сертификаты.md'},
    {'id': 'DOC-TU-22', 'doc_type': 'technical_conditions', 'number': '22.22.11-001-47645032-2022',
     'title': 'ТУ 22.22.11-001-47645032-2022 (термопакеты)', 'issuing_authority': 'ООО «ТЕРМИ»',
     'issued_date': '2022-01-01', 'expiry_date': None, 'status': 'valid',
     'file_path': None, 'public_allowed': True,
     'applicable_sku_ids': [], 'notes': None, 'source_reference': 'Патенты и сертификаты.md'},
    {'id': 'DOC-PAT-202340101', 'doc_type': 'patent', 'number': '202340101',
     'title': 'Патент (дизайн) № 202340101', 'issuing_authority': 'Роспатент',
     'issued_date': '2023-09-28', 'expiry_date': '2027-04-14', 'status': 'valid',
     'file_path': '/documents/202340101-PATENT-DES-2023-09-28.pdf', 'public_allowed': True,
     'applicable_sku_ids': [], 'notes': 'Действие с 15.04.2022 по 14.04.2027',
     'source_reference': 'Патенты и сертификаты.md'},
    {'id': 'DOC-EUR-PAT', 'doc_type': 'patent', 'number': 'EUR-202340101',
     'title': 'Европейский патент на ПО № 202340101', 'issuing_authority': 'EPO',
     'issued_date': None, 'expiry_date': None, 'status': 'valid',
     'file_path': '/documents/Evr_patent_na_PO_202340101.pdf', 'public_allowed': True,
     'applicable_sku_ids': [], 'notes': None, 'source_reference': 'Патенты и сертификаты.md'},
    {'id': 'DOC-PM-228420', 'doc_type': 'patent', 'number': '228420',
     'title': 'Свидетельство № 228420 на ПМ', 'issuing_authority': 'Роспатент',
     'issued_date': None, 'expiry_date': None, 'status': 'valid',
     'file_path': '/documents/Sv-vo_228420_na_PM.pdf', 'public_allowed': True,
     'applicable_sku_ids': [], 'notes': 'В архиве; уточнить статус у юриста.',
     'source_reference': 'Патенты и сертификаты.md'},
    {'id': 'DOC-PROTO-LAB', 'doc_type': 'test_protocol', 'number': 'LAB-001',
     'title': 'Лабораторный протокол испытаний (МГУПП, МГУТУ, ВНИИС)', 'issuing_authority': 'МГУПП / МГУТУ / ВНИИС',
     'issued_date': None, 'expiry_date': None, 'status': 'valid',
     'file_path': '/documents/Protokol_ispytanii.pdf', 'public_allowed': True,
     'applicable_sku_ids': [], 'notes': None, 'source_reference': 'Патенты и сертификаты.md'},
    {'id': 'DOC-TM-TERMY', 'doc_type': 'trademark', 'number': 'TM-TERMY',
     'title': 'Товарный знак *termy\\**', 'issuing_authority': 'Роспатент',
     'issued_date': '2022-11-03', 'expiry_date': None, 'status': 'valid',
     'file_path': None, 'public_allowed': True,
     'applicable_sku_ids': [], 'notes': 'Зарегистрирован 03.11.2022.',
     'source_reference': 'Патенты и сертификаты.md'},
    {'id': 'DOC-TM-SHOCK', 'doc_type': 'trademark', 'number': 'TM-SHOCK',
     'title': 'Товарный знак SHOCK®', 'issuing_authority': 'Роспатент',
     'issued_date': None, 'expiry_date': None, 'status': 'valid',
     'file_path': None, 'public_allowed': True,
     'applicable_sku_ids': [], 'notes': 'termy SHOCK® — для АХ.',
     'source_reference': 'Патенты и сертификаты.md'},
    {'id': 'DOC-TM-LINERBOX', 'doc_type': 'trademark', 'number': 'TM-LINERBOX',
     'title': 'Товарный знак LinerBox®', 'issuing_authority': 'Роспатент',
     'issued_date': None, 'expiry_date': None, 'status': 'valid',
     'file_path': None, 'public_allowed': True,
     'applicable_sku_ids': [], 'notes': 'ТЕРМИ LinerBox® — для термокоробов.',
     'source_reference': 'Патенты и сертификаты.md'},
    {'id': 'DOC-HALAL', 'doc_type': 'certificate', 'number': 'PENDING',
     'title': 'Сертификат «Халяль»', 'issuing_authority': None,
     'issued_date': None, 'expiry_date': None, 'status': 'in_renewal',
     'file_path': None, 'public_allowed': True,
     'applicable_sku_ids': [], 'notes': 'В процессе получения. Цель Q1 2026.',
     'source_reference': 'Патенты и сертификаты.md'},
    {'id': 'DOC-RU-SHOCK', 'doc_type': 'certificate', 'number': 'NOT_STARTED',
     'title': 'РУ Росздравнадзора для SHOCK®', 'issuing_authority': 'Росздравнадзор',
     'issued_date': None, 'expiry_date': None, 'status': 'not_started',
     'file_path': None, 'public_allowed': False,
     'applicable_sku_ids': [],
     'notes': 'Процедура не начата. Блокер для фармацевтических заявлений по SHOCK®.',
     'source_reference': 'Патенты и сертификаты.md'},
]
write_json('regulatory_documents.json', canonical)

# ===== Materials (hardcoded from vault notes) =====
materials = [
    {'id': 'MAT-MET', 'code': 'MET_85', 'name_ru': 'Металлизированная плёнка 85 мкм', 'function_in_product': 'layer_outer', 'nomenclature_1c': None, 'unit': 'м²'},
    {'id': 'MAT-VPE-500', 'code': 'VPE_500', 'name_ru': 'Вспененный полиэтилен 500 мкм', 'function_in_product': 'layer_middle', 'nomenclature_1c': None, 'unit': 'м²'},
    {'id': 'MAT-VPE-1000', 'code': 'VPE_1000', 'name_ru': 'Вспененный полиэтилен 1000 мкм', 'function_in_product': 'layer_middle', 'nomenclature_1c': None, 'unit': 'м²'},
    {'id': 'MAT-VPP-1CM', 'code': 'VPP_1CM', 'name_ru': 'Воздушно-пузырьковая плёнка 1 см (PRO)', 'function_in_product': 'layer_middle', 'nomenclature_1c': None, 'unit': 'м²'},
    {'id': 'MAT-VPP-2CM', 'code': 'VPP_2CM', 'name_ru': 'Воздушно-пузырьковая плёнка 2 см (LinerBox)', 'function_in_product': 'layer_middle', 'nomenclature_1c': None, 'unit': 'м²'},
    {'id': 'MAT-PVD', 'code': 'PVD_45', 'name_ru': 'ПВД 45 мкм', 'function_in_product': 'layer_inner', 'nomenclature_1c': None, 'unit': 'м²'},
    {'id': 'MAT-BOPP', 'code': 'BOPP', 'name_ru': 'БОПП-плёнка с ламинацией (термошопперы)', 'function_in_product': 'layer_outer', 'nomenclature_1c': None, 'unit': 'м²'},
    {'id': 'MAT-SPAN', 'code': 'SPAN_280', 'name_ru': 'Спанбонд 280 мкм (термошопперы)', 'function_in_product': 'layer_middle', 'nomenclature_1c': None, 'unit': 'м²'},
    {'id': 'MAT-T24', 'code': 'T24', 'name_ru': 'Гофрокартон Т24 (LinerBox)', 'function_in_product': 'frame', 'nomenclature_1c': None, 'unit': 'м²'},
    {'id': 'MAT-PLASTIC', 'code': 'PLASTIC_6MM', 'name_ru': 'Пищевой пластик 6 мм (дно)', 'function_in_product': 'bottom', 'nomenclature_1c': None, 'unit': 'шт'},
    {'id': 'MAT-GEL', 'code': 'GEL_SHOCK', 'name_ru': 'Гель SHOCK®', 'function_in_product': 'coolant', 'nomenclature_1c': None, 'unit': 'кг'},
]
write_json('materials.json', materials)

# ===== Carriers =====
carriers = [
    {'id': 'CR-TERMYMOBILE', 'name': 'ТермиМобиль', 'is_preferred_near_warehouse': True, 'covers_moscow': True, 'covers_russia': False, 'priority': 1, 'notes': 'Доставка по Москве за 24 ч'},
    {'id': 'CR-PEC', 'name': 'ПЭК', 'is_preferred_near_warehouse': True, 'covers_moscow': True, 'covers_russia': True, 'priority': 2, 'notes': None},
    {'id': 'CR-DL', 'name': 'Деловые Линии', 'is_preferred_near_warehouse': True, 'covers_moscow': True, 'covers_russia': True, 'priority': 3, 'notes': None},
    {'id': 'CR-ENERGY', 'name': 'Энергия', 'is_preferred_near_warehouse': True, 'covers_moscow': True, 'covers_russia': True, 'priority': 4, 'notes': None},
    {'id': 'CR-KIT', 'name': 'КИТ', 'is_preferred_near_warehouse': True, 'covers_moscow': True, 'covers_russia': True, 'priority': 5, 'notes': None},
    {'id': 'CR-VGS', 'name': 'Все грузы Севера', 'is_preferred_near_warehouse': True, 'covers_moscow': True, 'covers_russia': True, 'priority': 6, 'notes': None},
    {'id': 'CR-RAILKONT', 'name': 'Рейл Континент', 'is_preferred_near_warehouse': True, 'covers_moscow': True, 'covers_russia': True, 'priority': 7, 'notes': None},
    {'id': 'CR-BAIKAL', 'name': 'Байкал Сервис', 'is_preferred_near_warehouse': True, 'covers_moscow': True, 'covers_russia': True, 'priority': 8, 'notes': None},
    {'id': 'CR-MAGIC', 'name': 'Мейджик Транс', 'is_preferred_near_warehouse': True, 'covers_moscow': True, 'covers_russia': True, 'priority': 9, 'notes': None},
    {'id': 'CR-UTEK', 'name': 'УТЭК', 'is_preferred_near_warehouse': True, 'covers_moscow': True, 'covers_russia': True, 'priority': 10, 'notes': None},
    {'id': 'CR-STHOLD', 'name': 'СТ-Холдинг', 'is_preferred_near_warehouse': True, 'covers_moscow': True, 'covers_russia': True, 'priority': 11, 'notes': None},
    {'id': 'CR-CDEK', 'name': 'СДЭК', 'is_preferred_near_warehouse': True, 'covers_moscow': True, 'covers_russia': True, 'priority': 12, 'notes': 'Используется для отправки образцов.'},
]
write_json('carriers.json', carriers)

# ===== Customer segments (13) =====
segments = [
    {'id': 'SEG-DELIVERY', 'name': 'Службы доставки', 'segment_type': 'b2b_termopakety',
     'decision_maker': 'Менеджер по закупкам / логистике, операционный директор, владелец',
     'pains': 'Ненадёжность упаковки, жалобы клиентов, дефицит в пиковый сезон',
     'expectations': 'Прочность, надёжные швы, складские запасы, быстрая отгрузка',
     'offer': 'Сбермаркет использует до 30 циклов; оперативная замена при рекламации', 'constraints': None},
    {'id': 'SEG-RETAIL', 'name': 'Крупный ритейл', 'segment_type': 'b2b_termopakety',
     'decision_maker': 'Категорийный закупщик',
     'pains': 'Цена закупки, невыполнение KPI, качество печати',
     'expectations': 'Низкая цена, флексопечать по Pantone, сертификаты, стабильность',
     'offer': 'Маркетинговые материалы, комплект пакет + АКХ', 'constraints': None},
    {'id': 'SEG-DIST', 'name': 'Дистрибьюторы упаковки', 'segment_type': 'distributor',
     'decision_maker': 'Менеджер по закупкам',
     'pains': 'Маржинальность, срывы поставок',
     'expectations': 'Конкурентная цена, ассортимент, сроки',
     'offer': '3 ценовых сегмента, резерв + 24 ч, спеццены на первые 3 заказа', 'constraints': None},
    {'id': 'SEG-PROD', 'name': 'Производители скоропорта', 'segment_type': 'b2b_termopakety',
     'decision_maker': 'Снабжение / логистика / технолог',
     'pains': 'Порча товара → убытки и репутационные потери',
     'expectations': 'Инженерно просчитанное решение, экспертный подбор',
     'offer': 'Уникальный формат 42×50; оперативная реакция', 'constraints': None},
    {'id': 'SEG-PHARMA', 'name': 'Фармацевтика / аптеки / «Красота и здоровье»', 'segment_type': 'pharma',
     'decision_maker': 'Руководитель логистики / спец.по качеству / закупки',
     'pains': 'Нарушение холодовой цепи, compliance-риски',
     'expectations': 'Документированное соответствие, протоколы, стабильность',
     'offer': 'Протоколы испытаний',
     'constraints': 'РУ Росздравнадзора не получено — без фарм. регуляторных заявлений'},
    {'id': 'SEG-HORECA', 'name': 'HoReCa (рестораны, кафе, кофейни, кейтеринг)', 'segment_type': 'b2b_shoppers',
     'decision_maker': 'Владелец, управляющий, шеф-повар, закупщик',
     'pains': 'Остывшая еда, плохой имидж, стоимость одноразовых решений',
     'expectations': 'Термозащита до 2 ч, дизайн, логотип, гибкий МОЗ',
     'offer': 'Термошопперы Termy от 100 шт', 'constraints': None},
    {'id': 'SEG-QCOMM', 'name': 'Dark Store / Q-Commerce', 'segment_type': 'b2b_shoppers',
     'decision_maker': 'Основатель, операционный директор, маркетинг',
     'pains': 'Потеря качества, дешёвый образ',
     'expectations': 'Технологичность, устойчивое дно, СТМ-печать (QR)',
     'offer': 'Пробная партия + короткие сроки', 'constraints': None},
    {'id': 'SEG-RETAIL-SHOP', 'name': 'Ритейл (шопперы для перепродажи)', 'segment_type': 'b2b_shoppers',
     'decision_maker': 'Категорийный закупщик',
     'pains': 'Дополнительный сервис для клиента, лояльность',
     'expectations': 'Крупный опт, сертификаты, поставки на РЦ/ТТ',
     'offer': 'Скидки за объём', 'constraints': None},
    {'id': 'SEG-APTEKA', 'name': 'Аптеки / бьюти / клиники', 'segment_type': 'b2b_shoppers',
     'decision_maker': 'Закупщик / маркетинг / снабженец',
     'pains': 'Нарушение температурного режима, конфиденциальность',
     'expectations': 'Непрозрачные материалы, клапан-скотч, ТР ТС 005/2011',
     'offer': 'Декларация ТР ТС + клапан-скотч', 'constraints': None},
    {'id': 'SEG-AGENCY', 'name': 'Рекламные агентства', 'segment_type': 'b2b_shoppers',
     'decision_maker': 'Маркетологи / закупки',
     'pains': 'Бесполезный промо, низкий ROI',
     'expectations': 'Функциональность + стиль, СТМ, малые тиражи',
     'offer': 'Бесплатный макет, шелко от 100 шт', 'constraints': None},
    {'id': 'SEG-B2C-PRO', 'name': 'B2C: Городские профессионалы', 'segment_type': 'b2c',
     'decision_maker': 'Конечный потребитель 28–40',
     'pains': 'Некрасивый термос, неудобство, ненадёжность',
     'expectations': 'Стиль + технологичность', 'offer': 'Термокружки termy* 600/900 мл',
     'constraints': 'Направление сворачивается'},
    {'id': 'SEG-B2C-TRAVEL', 'name': 'B2C: Путешественники/авто', 'segment_type': 'b2c',
     'decision_maker': 'Конечный потребитель 25–45',
     'pains': 'Громоздкость, хрупкость, мытьё',
     'expectations': 'Надёжность, объём', 'offer': 'Термокружки 900 мл', 'constraints': None},
    {'id': 'SEG-B2C-ACTIVE', 'name': 'B2C: Активный отдых, семьи с детьми', 'segment_type': 'b2c',
     'decision_maker': 'Конечный потребитель',
     'pains': 'Хранение в дороге', 'expectations': 'Объём, надёжность',
     'offer': 'Термошопперы в рознице', 'constraints': None},
]
write_json('customer_segments.json', segments)

# ===== CRM statuses =====
crm_statuses = [
    {'id': 'CRM-ACT', 'code': 'active', 'name_ru': 'Действующий', 'description': 'Хотя бы один раз оплатил и купил. Поддержание и развитие.'},
    {'id': 'CRM-WRK', 'code': 'in_work', 'name_ru': 'В работе', 'description': 'Потенциальный клиент (лид), в переговорах.'},
    {'id': 'CRM-REJ', 'code': 'rejected', 'name_ru': 'Отказ', 'description': 'Отказ от сотрудничества.'},
    {'id': 'CRM-MAIL', 'code': 'in_mailing', 'name_ru': 'В рассылке', 'description': 'Периодические касания.'},
]
write_json('crm_statuses.json', crm_statuses)

# ===== Legal entities TERMY =====
legal_entities = [
    {'id': 'LE-TERMY', 'name': 'ООО «ТЕРМИ»', 'inn': '1227700339976', 'vat_profile': 'with_vat', 'role': 'main'},
    {'id': 'LE-BIGUPACK', 'name': 'ООО «Бигупак»', 'inn': None, 'vat_profile': 'with_vat', 'role': 'additional'},
    {'id': 'LE-IP-SULTAN', 'name': 'ИП Султанов М. Р.', 'inn': None, 'vat_profile': 'without_vat', 'role': 'distributor'},
]
write_json('legal_entities_termy.json', legal_entities)

# ===== Printing partners =====
printing_partners = [
    {'id': 'PP-LAIR', 'name': 'Лаир', 'specialization': 'silkscreen',
     'stability_note': 'Нестабильность цветов (красный → бордовый, белый полупрозрачный). Идёт поиск альтернативы.',
     'contact_info': None},
    {'id': 'PP-PARTNERMEDIA', 'name': 'Партнёр-Медиа', 'specialization': 'flexo',
     'stability_note': 'Технолог партнёра согласует более мелкие кегли.',
     'contact_info': None},
]
write_json('printing_partners.json', printing_partners)

# ===== Roles =====
roles = [
    {'id': 'R-ADMIN', 'code': 'admin', 'name_ru': 'Администратор системы', 'confirmed': True, 'pending_validation': False,
     'description': 'Полный доступ. Управление пользователями, ролями, справочниками, audit log.'},
    {'id': 'R-DIRECTOR', 'code': 'director', 'name_ru': 'Генеральный директор', 'confirmed': True, 'pending_validation': False,
     'description': 'Право подписи по Уставу. Просмотр всех разделов.'},
    {'id': 'R-ROP', 'code': 'rop', 'name_ru': 'РОП (руководитель отдела продаж)', 'confirmed': True, 'pending_validation': False,
     'description': 'Согласование цен, СТМ, договоров. Одобрение крупных наборов образцов.'},
    {'id': 'R-MANAGER', 'code': 'manager', 'name_ru': 'Менеджер активных продаж', 'confirmed': True, 'pending_validation': False,
     'description': 'Работа с клиентом, КП, заказ, макет, образцы, NPS.'},
    {'id': 'R-ACCOUNTANT', 'code': 'accountant', 'name_ru': 'Бухгалтерия', 'confirmed': True, 'pending_validation': False,
     'description': 'УПД, акты сверок, декларации.'},
    {'id': 'R-SECRETARY', 'code': 'secretary', 'name_ru': 'Секретарь', 'confirmed': True, 'pending_validation': False,
     'description': 'Почта России, несрочная документация.'},
    {'id': 'R-DESIGNER', 'code': 'designer', 'name_ru': 'Дизайнер макетов', 'confirmed': True, 'pending_validation': False,
     'description': 'Работа с макетами СТМ, чек-лист валидации.'},
    {'id': 'R-DRIVER', 'code': 'driver', 'name_ru': 'Водитель ТермиМобиль', 'confirmed': True, 'pending_validation': False,
     'description': 'Доставка по Москве.'},
    {'id': 'R-PROD-HEAD', 'code': 'prod_head_inferred', 'name_ru': 'Начальник производства (pending validation)',
     'confirmed': False, 'pending_validation': True,
     'description': 'Валидация на визите (TERMY_04A).'},
    {'id': 'R-PROD-MASTER', 'code': 'prod_master_inferred', 'name_ru': 'Мастер/бригадир (pending validation)',
     'confirmed': False, 'pending_validation': True, 'description': 'Валидация на визите.'},
    {'id': 'R-PROD-OP', 'code': 'prod_operator_inferred', 'name_ru': 'Оператор участка (pending validation)',
     'confirmed': False, 'pending_validation': True, 'description': 'Валидация на визите.'},
    {'id': 'R-PROD-PACK', 'code': 'prod_packer_inferred', 'name_ru': 'Упаковщик (pending validation)',
     'confirmed': False, 'pending_validation': True, 'description': 'Валидация на визите.'},
    {'id': 'R-WAREHOUSE', 'code': 'warehouseman_inferred', 'name_ru': 'Кладовщик (pending validation)',
     'confirmed': False, 'pending_validation': True, 'description': 'Валидация на визите.'},
    {'id': 'R-QC', 'code': 'qc_inferred', 'name_ru': 'Контролёр качества (pending validation)',
     'confirmed': False, 'pending_validation': True, 'description': 'Валидация на визите.'},
]
write_json('roles.json', roles)

# ===== sku_document_map =====
# Blanket: привязать все основные ЕАЭС-декларации ко всем термопакетам/LinerBox/BubblePak/шопперам
doc_map = []
for sku in skus:
    pl_id = sku.get('product_line_id')
    if pl_id in ('PL-001', 'PL-002', 'PL-003', 'PL-004', 'PL-005', 'PL-006', 'PL-007', 'PL-009'):
        for doc_id in ('DOC-DEC-902483', 'DOC-DEC-1158277', 'DOC-DEC-1158278', 'DOC-TU-22', 'DOC-ISO-9001', 'DOC-PROTO-LAB'):
            doc_map.append({'sku_id': sku['id'], 'document_id': doc_id, 'scope_note': None})
    # SHOCK привязки — ТУ + ISO + протокол, без декларации (SHOCK — не тара)
    if pl_id in ('PL-008', 'PL-011'):
        for doc_id in ('DOC-ISO-9001', 'DOC-PROTO-LAB'):
            doc_map.append({'sku_id': sku['id'], 'document_id': doc_id, 'scope_note': None})
        # Отдельно помечаем отсутствие РУ
        doc_map.append({'sku_id': sku['id'], 'document_id': 'DOC-RU-SHOCK', 'scope_note': 'not_started'})
    # Patent на ручку-защёлку — Термопакеты (PL-001..004)
    if pl_id in ('PL-001', 'PL-002', 'PL-003'):
        doc_map.append({'sku_id': sku['id'], 'document_id': 'DOC-PAT-202340101', 'scope_note': None})
    # Trademark LinerBox
    if pl_id == 'PL-007':
        doc_map.append({'sku_id': sku['id'], 'document_id': 'DOC-TM-LINERBOX', 'scope_note': None})
    # Trademark SHOCK
    if pl_id in ('PL-008', 'PL-011'):
        doc_map.append({'sku_id': sku['id'], 'document_id': 'DOC-TM-SHOCK', 'scope_note': None})
    # Trademark termy
    doc_map.append({'sku_id': sku['id'], 'document_id': 'DOC-TM-TERMY', 'scope_note': None})
write_json('sku_document_map.json', doc_map)

# ===== sku_printing_support =====
# По правилу: PL-001..005 и PL-007 поддерживают шелко и флексо.
# Zip-Lock (PL-003) — флексо от 22 000. Клапан-скотч (PL-004) — флексо от 25 000. БаблПак (PL-005) — флексо от 44 000.
# PL-006 (Термошопперы) — шелко. PL-008 SHOCK — этикетка / выбит в пластике (moq в overrides).
# PL-010 термокружки — гравировка СТМ (из коробки 36 шт).
pm_silk = 'PM-01'
pm_flexo = 'PM-02'
support = []
for sku in skus:
    pl = sku['product_line_id']
    sid = sku['id']
    if pl in ('PL-001', 'PL-002'):
        support.append({'sku_id': sid, 'printing_method_id': pm_silk, 'moq_override': None, 'lead_time_override_weeks': None, 'notes': None})
        support.append({'sku_id': sid, 'printing_method_id': pm_flexo, 'moq_override': None, 'lead_time_override_weeks': None, 'notes': None})
    elif pl == 'PL-003':
        support.append({'sku_id': sid, 'printing_method_id': pm_flexo, 'moq_override': 22000, 'lead_time_override_weeks': '4–6', 'notes': 'Только на мет.плёнке.'})
    elif pl == 'PL-004':
        support.append({'sku_id': sid, 'printing_method_id': pm_flexo, 'moq_override': 25000, 'lead_time_override_weeks': '4–5', 'notes': 'Только на мет.плёнке.'})
    elif pl == 'PL-005':
        support.append({'sku_id': sid, 'printing_method_id': pm_flexo, 'moq_override': 44000, 'lead_time_override_weeks': None, 'notes': None})
    elif pl == 'PL-007':
        support.append({'sku_id': sid, 'printing_method_id': pm_flexo, 'moq_override': None, 'lead_time_override_weeks': None, 'notes': 'Печать логотипа на коробе.'})
    elif pl == 'PL-006':
        support.append({'sku_id': sid, 'printing_method_id': pm_silk, 'moq_override': 100, 'lead_time_override_weeks': None, 'notes': 'Бесплатный макет.'})
    elif pl == 'PL-008':
        support.append({'sku_id': sid, 'printing_method_id': pm_silk, 'moq_override': 1000, 'lead_time_override_weeks': '2–3', 'notes': 'Этикетка с лого клиента.'})
    elif pl == 'PL-010':
        support.append({'sku_id': sid, 'printing_method_id': pm_silk, 'moq_override': 36, 'lead_time_override_weeks': '2', 'notes': 'Гравировка СТМ, в нашем цвете.'})
write_json('sku_printing_support.json', support)

# ===== sku_materials (generic rough mapping — refined later) =====
sku_mat = []
for sku in skus:
    pl = sku['product_line_id']
    sid = sku['id']
    if pl == 'PL-001':
        for m in ('MAT-MET', 'MAT-VPE-1000', 'MAT-PVD'):
            sku_mat.append({'sku_id': sid, 'material_id': m, 'function': 'layer_outer' if m == 'MAT-MET' else 'layer_middle' if 'VPE' in m else 'layer_inner',
                             'consumption_norm_per_unit': None, 'tolerance_pct': None})
    elif pl == 'PL-002':
        for m, f in (('MAT-MET', 'layer_outer'), ('MAT-VPE-1000', 'layer_middle'), ('MAT-MET', 'layer_inner'), ('MAT-PLASTIC', 'bottom')):
            sku_mat.append({'sku_id': sid, 'material_id': m, 'function': f, 'consumption_norm_per_unit': None, 'tolerance_pct': None})
    elif pl == 'PL-003':
        sku_mat.append({'sku_id': sid, 'material_id': 'MAT-MET', 'function': 'layer_outer', 'consumption_norm_per_unit': None, 'tolerance_pct': None})
    elif pl == 'PL-004':
        for m, f in (('MAT-MET', 'layer_outer'), ('MAT-PVD', 'layer_inner')):
            sku_mat.append({'sku_id': sid, 'material_id': m, 'function': f, 'consumption_norm_per_unit': None, 'tolerance_pct': None})
    elif pl == 'PL-005':
        for m, f in (('MAT-VPP-1CM', 'layer_middle'), ('MAT-MET', 'layer_outer')):
            sku_mat.append({'sku_id': sid, 'material_id': m, 'function': f, 'consumption_norm_per_unit': None, 'tolerance_pct': None})
    elif pl == 'PL-006':
        for m, f in (('MAT-BOPP', 'layer_outer'), ('MAT-SPAN', 'layer_middle'), ('MAT-BOPP', 'layer_inner')):
            sku_mat.append({'sku_id': sid, 'material_id': m, 'function': f, 'consumption_norm_per_unit': None, 'tolerance_pct': None})
    elif pl == 'PL-007':
        for m, f in (('MAT-T24', 'frame'), ('MAT-MET', 'layer_outer'), ('MAT-VPP-2CM', 'layer_middle'), ('MAT-PVD', 'layer_inner')):
            sku_mat.append({'sku_id': sid, 'material_id': m, 'function': f, 'consumption_norm_per_unit': None, 'tolerance_pct': None})
    elif pl == 'PL-008':
        for m, f in (('MAT-PLASTIC', 'frame'), ('MAT-GEL', 'coolant')):
            sku_mat.append({'sku_id': sid, 'material_id': m, 'function': f, 'consumption_norm_per_unit': None, 'tolerance_pct': None})
write_json('sku_materials.json', sku_mat)

print('\nDONE. Real seed generated to', OUT)
